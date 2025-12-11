import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Đọc file test_cases.md và parse test cases
test_cases_data = []

# Module A: Authentication (11 cases)
test_cases_data.extend([
    ["A.1", "Đăng ký khách hàng qua OTP", "Authentication", "P0", "Database có sẵn, email service hoạt động", "1. Truy cập trang đăng ký\n2. Nhập email & phone\n3. Nhấn 'Gửi OTP'\n4. Nhận OTP từ email\n5. Nhập OTP\n6. Điền username, password, họ tên\n7. Nhấn 'Đăng ký'", "OTP được gửi thành công\nVerify OTP thành công\nTài khoản được tạo\nRedirect về trang đăng nhập"],
    ["A.2", "Đăng ký nhân viên qua OTP", "Authentication", "P0", "Database có sẵn, email service hoạt động", "1. Truy cập trang đăng ký NV\n2. Nhập email & phone\n3. Gửi OTP và verify\n4. Điền thông tin + upload giấy tờ\n5. Nhấn 'Đăng ký'", "Tài khoản NV được tạo với trạng thái 'pending'\nAdmin nhận được yêu cầu duyệt"],
    ["A.3", "Đăng nhập thành công", "Authentication", "P0", "Có tài khoản hợp lệ", "1. Truy cập trang đăng nhập\n2. Nhập username\n3. Nhập password\n4. Nhấn 'Đăng nhập'", "Đăng nhập thành công\nRedirect về trang chủ\nToken được lưu"],
    ["A.4", "Đăng nhập thất bại - sai mật khẩu", "Authentication", "P0", "Có tài khoản hợp lệ", "1. Truy cập trang đăng nhập\n2. Nhập username đúng\n3. Nhập password SAI\n4. Nhấn 'Đăng nhập'", "Hiển thị lỗi 'Tên đăng nhập hoặc mật khẩu không đúng'\nKhông đăng nhập được"],
    ["A.5", "Đăng nhập thất bại - user không tồn tại", "Authentication", "P0", "Database có sẵn", "1. Nhập username không tồn tại\n2. Nhập password bất kỳ\n3. Nhấn 'Đăng nhập'", "Hiển thị lỗi 'Tên đăng nhập hoặc mật khẩu không đúng'"],
    ["A.6", "Quên mật khẩu qua OTP", "Authentication", "P0", "Có tài khoản hợp lệ", "1. Nhấn 'Quên mật khẩu'\n2. Nhập email/phone\n3. Gửi OTP\n4. Nhập OTP\n5. Nhập mật khẩu mới\n6. Xác nhận", "OTP được gửi\nĐổi mật khẩu thành công\nĐăng nhập được bằng mật khẩu mới"],
    ["A.7", "Đổi mật khẩu khi đã đăng nhập", "Authentication", "P1", "Đã đăng nhập", "1. Vào trang Profile\n2. Nhấn 'Đổi mật khẩu'\n3. Nhập mật khẩu cũ\n4. Nhập mật khẩu mới\n5. Xác nhận mật khẩu mới\n6. Nhấn 'Lưu'", "Đổi mật khẩu thành công\nĐăng nhập được bằng mật khẩu mới"],
    ["A.8", "Logout", "Authentication", "P0", "Đã đăng nhập", "1. Nhấn nút 'Đăng xuất'", "Token bị xóa\nRedirect về trang đăng nhập\nKhông truy cập được protected routes"],
    ["A.9", "Truy cập không có quyền - khách hàng vào admin", "Authorization", "P0", "Đăng nhập với tài khoản khách hàng", "1. Truy cập URL /admin/dashboard", "Hiển thị lỗi 403 Forbidden hoặc redirect về trang chủ"],
    ["A.10", "Token hết hạn", "Authentication", "P0", "Đã đăng nhập, token hết hạn", "1. Đợi token hết hạn\n2. Gọi API protected", "API trả về 401 Unauthorized\nYêu cầu đăng nhập lại"],
    ["A.11", "Rate limit OTP", "Authentication", "P1", "Email service hoạt động", "1. Gửi OTP lần 1\n2. Gửi OTP lần 2 ngay lập tức\n3. Gửi OTP lần 3 ngay lập tức", "Sau 3-5 lần → Hiển thị 'Bạn đã gửi quá nhiều yêu cầu, vui lòng thử lại sau'"],
])

# Module B: Profile (7 cases)
test_cases_data.extend([
    ["B.1", "Xem profile", "Profile", "P1", "Đã đăng nhập", "1. Vào trang Profile\n2. API: GET /api/auth/profile", "Hiển thị đúng thông tin: username, tên, email, phone"],
    ["B.2", "Cập nhật profile thành công", "Profile", "P1", "Đã đăng nhập", "1. Vào trang Profile\n2. Sửa tên, email\n3. Nhấn 'Cập nhật'\n4. API: PUT /api/auth/profile", "Cập nhật thành công\nHiển thị thông báo 'Cập nhật thành công'\nThông tin mới được lưu"],
    ["B.3", "Upload avatar thành công", "Profile", "P1", "Đã đăng nhập", "1. Nhấn 'Thay đổi ảnh đại diện'\n2. Chọn file ảnh (jpg/png) < 2MB\n3. Upload", "Upload thành công\nẢnh mới hiển thị\nURL ảnh được lưu trong database"],
    ["B.4", "Upload avatar sai định dạng", "Profile", "P1", "Đã đăng nhập", "1. Chọn file .exe hoặc .txt\n2. Upload", "Hiển thị lỗi 'Chỉ chấp nhận ảnh jpg, png'\nKhông upload được"],
    ["B.5", "Upload avatar quá dung lượng", "Profile", "P1", "Đã đăng nhập", "1. Chọn file ảnh > 2MB\n2. Upload", "Hiển thị lỗi 'Kích thước file > 2MB'\nKhông upload được"],
    ["B.6", "Cập nhật lịch làm việc nhân viên", "Profile", "P1", "Đăng nhập với tài khoản nhân viên", "1. Vào 'Lịch làm việc'\n2. Chọn thứ 2-6, 8:00-17:00\n3. Lưu\n4. API: PUT /api/staff/schedules", "Lịch được cập nhật\nHiển thị đúng trên calendar"],
    ["B.7", "Admin duyệt nhân viên", "Admin", "P0", "NV đã đăng ký, upload giấy tờ", "1. Admin login\n2. Vào 'Quản lý ứng viên'\n3. Xem hồ sơ NV\n4. Nhấn 'Duyệt'", "Trạng thái NV chuyển thành 'active'\nNV có thể nhận đơn"],
])

# Module C: Address (5 cases)
test_cases_data.extend([
    ["C.1", "Xem danh sách địa chỉ", "Address", "P1", "Đã đăng nhập, có địa chỉ đã lưu", "1. Vào 'Địa chỉ của tôi'\n2. API: GET /api/addresses", "Hiển thị danh sách địa chỉ đã lưu"],
    ["C.2", "Tạo địa chỉ mới thành công", "Address", "P1", "Đã đăng nhập", "1. Nhấn 'Thêm địa chỉ'\n2. Nhập địa chỉ: '140 Lê Trọng Tấn, Tân Phú'\n3. Nhập căn hộ: 'Căn 101'\n4. Nhấn 'Lưu'\n5. API: POST /api/addresses", "Địa chỉ được tạo\nTự động xác định quận 'Quận Tân Phú'\nHiển thị trong danh sách"],
    ["C.3", "Cập nhật địa chỉ", "Address", "P1", "Đã có địa chỉ", "1. Nhấn 'Sửa' trên địa chỉ\n2. Sửa căn hộ thành 'Căn 102'\n3. Nhấn 'Lưu'\n4. API: PUT /api/addresses/{id}", "Địa chỉ được cập nhật\nThông tin mới hiển thị"],
    ["C.4", "Xóa địa chỉ (soft delete)", "Address", "P1", "Đã có địa chỉ", "1. Nhấn 'Xóa' trên địa chỉ\n2. Xác nhận xóa\n3. API: DELETE /api/addresses/{id}", "Địa chỉ không hiển thị nữa\nTrong database: is_deleted = true"],
    ["C.5", "Validation - địa chỉ rỗng", "Address", "P1", "Đã đăng nhập", "1. Nhấn 'Thêm địa chỉ'\n2. Để trống\n3. Nhấn 'Lưu'", "Hiển thị lỗi 'Vui lòng nhập địa chỉ'"],
])

# Module D: Customer Booking (13 cases)
test_cases_data.extend([
    ["D.1", "Đặt lịch theo giờ - cơ bản", "Booking", "P0", "Đã đăng nhập, có địa chỉ", "1. Vào 'Đặt lịch'\n2. Chọn 'Theo giờ'\n3. Chọn gói 2 giờ\n4. Chọn ngày mai\n5. Chọn giờ 9:00\n6. Chọn địa chỉ\n7. Chọn COD\n8. Nhấn 'Đặt lịch'", "Đơn được tạo với trạng thái finding_staff\nHiển thị 'Đặt lịch thành công'"],
    ["D.2", "Đặt lịch theo giờ - có voucher", "Booking", "P0", "Đã đăng nhập, có voucher hợp lệ", "1. Làm như D.1\n2. Tại màn thanh toán, nhập mã voucher\n3. Nhấn 'Áp dụng'\n4. Kiểm tra giá\n5. Đặt lịch", "Voucher được áp dụng\nGiá giảm đúng %\nĐơn được tạo với giá đã giảm"],
    ["D.3", "Đặt lịch giờ cao điểm - tự động thêm phụ thu", "Booking", "P0", "Đã đăng nhập", "1. Chọn gói 2 giờ\n2. Chọn giờ 7:00 (trước 8h)\n3. Kiểm tra giá", "Phụ thu PT001 (giờ cao điểm) được tự động thêm\nTổng tiền = giá gói + phụ thu"],
    ["D.4", "Đặt lịch cuối tuần - tự động thêm phụ thu", "Booking", "P0", "Đã đăng nhập", "1. Chọn gói 2 giờ\n2. Chọn ngày Thứ 7 hoặc CN\n3. Kiểm tra giá", "Phụ thu PT003 (cuối tuần) được tự động thêm\nTổng tiền = giá gói + phụ thu"],
    ["D.5", "Đặt lịch có thú cưng", "Booking", "P1", "Đã đăng nhập", "1. Chọn gói 2 giờ\n2. Tick 'Có thú cưng'\n3. Kiểm tra giá", "Phụ thu thú cưng được thêm\nTổng tiền tăng"],
    ["D.6", "Đặt lịch theo tháng - 20 buổi", "Booking", "P0", "Đã đăng nhập", "1. Chọn 'Theo tháng'\n2. Chọn gói 20 buổi\n3. Chọn lịch: Thứ 2,4,6 - 8:00\n4. Chọn ngày bắt đầu\n5. Thanh toán\n6. Đặt lịch", "Đơn tháng được tạo\n20 buổi được tạo tự động\nTrạng thái finding_staff"],
    ["D.7", "Validation - đặt ngày trong quá khứ", "Booking", "P0", "Đã đăng nhập", "1. Chọn ngày hôm qua\n2. Nhấn 'Tiếp tục'", "Hiển thị lỗi 'Không thể chọn ngày trong quá khứ'"],
    ["D.8", "Validation - đặt quá gần", "Booking", "P1", "Đã đăng nhập", "1. Chọn ngày hôm nay\n2. Chọn giờ 1 tiếng sau\n3. Nhấn 'Tiếp tục'", "Hiển thị cảnh báo 'Nên đặt trước ít nhất 2 giờ'"],
    ["D.9", "Tính giá động", "Booking", "P0", "Đã đăng nhập", "1. API: POST /booking/quote-hour\n2. Body: {duration: 2}", "Response trả về giá đúng của gói 2 giờ"],
    ["D.10", "Áp dụng nhiều voucher", "Booking", "P1", "Có 2 voucher", "1. Nhập voucher 1 → áp dụng\n2. Nhập voucher 2 → áp dụng", "Chỉ voucher cuối được áp dụng\nHoặc hiển thị 'Chỉ được dùng 1 voucher'"],
    ["D.11", "Voucher hết hạn", "Booking", "P1", "Có voucher hết hạn", "1. Nhập mã voucher hết hạn\n2. Nhấn 'Áp dụng'", "Hiển thị lỗi 'Voucher đã hết hạn'"],
    ["D.12", "Voucher đã dùng", "Booking", "P1", "Đã dùng voucher 'chỉ dùng 1 lần'", "1. Nhập lại mã voucher\n2. Nhấn 'Áp dụng'", "Hiển thị lỗi 'Bạn đã sử dụng voucher này'"],
    ["D.13", "Tìm nhân viên available", "Booking", "P0", "Có NV active", "1. API: POST /booking/find-staff\n2. Body: ngày, giờ, địa chỉ", "Trả về danh sách NV phù hợp với score, rating"],
])

# Module E: Staff Booking (10 cases)
test_cases_data.extend([
    ["E.1", "Xem danh sách đơn available", "Staff", "P0", "Đăng nhập NV, có đơn finding_staff", "1. NV login app\n2. Vào 'Đơn khả dụng'\n3. API: GET /api/staff/bookings/available", "Hiển thị danh sách đơn chưa có NV"],
    ["E.2", "Claim đơn giờ thành công", "Staff", "P0", "Có đơn available", "1. NV nhấn 'Nhận đơn' trên 1 đơn\n2. API: POST /api/staff/bookings/{id}/claim", "Đơn chuyển sang assigned\nNV được gán (ID_NV)\nKhách nhận thông báo"],
    ["E.3", "Confirm đơn đã được admin gán", "Staff", "P0", "Adminđã gán NV", "1. NV nhận thông báo có đơn mới\n2. Vào 'Đơn của tôi'\n3. Nhấn 'Xác nhận'\n4. API: POST /api/staff/bookings/{id}/confirm", "Đơn chuyển sang confirmed"],
    ["E.4", "Reject đơn", "Staff", "P0", "NV đã được gán đơn", "1. NV vào 'Đơn của tôi'\n2. Nhấn 'Từ chối'\n3. Nhập lý do\n4. Xác nhận\n5. API: POST /api/staff/bookings/{id}/reject", "Đơn quay về finding_staff\nID_NV = NULL\nKhách nhận thông báo"],
    ["E.5", "Nhiều NV claim cùng 1 đơn", "Staff", "P0", "2 NV cùng online, 1 đơn available", "1. NV1 nhấn 'Nhận đơn' → API claim\n2. NV2 nhấn 'Nhận đơn' → API claim (gần như đồng thời)", "NV1 claim thành công\nNV2 nhận lỗi 'Đơn đã có người nhận'"],
    ["E.6", "Claim buổi trong đơn tháng", "Staff", "P0", "Có đơn tháng, buổi finding_staff", "1. NV vào 'Đơn tháng khả dụng'\n2. Chọn 1 buổi\n3. Nhấn 'Nhận'\n4. API: POST /api/staff/month-sessions/{id}/claim", "Buổi đó được gán NV\nTrạng thái assigned"],
    ["E.7", "Hoàn thành đơn", "Staff", "P0", "Đơn confirmed", "1. NV vào 'Đơn của tôi'\n2. Nhấn 'Hoàn thành'\n3. API: POST /api/staff/bookings/{id}/complete", "Đơn chuyển completed\nTiền được cộng vào ví NV\nKhách nhận thông báo đánh giá"],
    ["E.8", "Kiểm tra lịch chồng lấn", "Staff", "P0", "NV đã có đơn 9:00-11:00 ngày mai", "1. NV claim đơn khác 10:00-12:00 ngày mai", "Hiển thị lỗi 'Bạn đã có lịch trong khung giờ này'"],
    ["E.9", "Xem lịch làm việc", "Staff", "P1", "Đăng nhập NV", "1. Vào 'Lịch làm việc'\n2. API: GET /api/staff/schedules", "Hiển thị lịch theo tuần: Thứ, giờ bắt đầu, giờ kết thúc"],
    ["E.10", "Cập nhật lịch làm việc", "Staff", "P1", "Đăng nhập NV", "1. Sửa lịch: Thứ 2-6, 8:00-18:00\n2. Lưu\n3. API: PUT /api/staff/schedules", "Lịch được cập nhật thành công"],
])

# Module F: Payment & Refund (9 cases)
test_cases_data.extend([
    ["F.1", "Thanh toán COD", "Payment", "P0", "Đã đăng nhập", "1. Đặt lịch\n2. Chọn 'Thanh toán khi hoàn thành'\n3. Xác nhận", "Đơn được tạo với trạng thái finding_staff\nKhông có bản ghi thanh toán trong LichSuThanhToan"],
    ["F.2", "Thanh toán VNPay thành công", "Payment", "P0", "VNPay sandbox hoạt động", "1. Đặt lịch\n2. Chọn 'VNPay'\n3. Redirect sang VNPay\n4. Thanh toán với card test\n5. Callback về hệ thống", "Thanh toán thành công\nĐơn finding_staff\nBản ghi trong LichSuThanhToan với LoaiGD = payment"],
    ["F.3", "Thanh toán VNPay thất bại", "Payment", "P0", "VNPay sandbox", "1. Đặt lịch\n2. Chọn VNPay\n3. Hủy thanh toán trên VNPay", "Quay về trang đặt lịch\nĐơn không được tạo hoặc cancelled"],
    ["F.4", "Webhook VNPay trễ", "Payment", "P1", "VNPay sandbox", "1. Thanh toán VN Pay\n2. Giả lập webhook đến sau 5 phút\n3. Server nhận webhook", "Trạng thái đơn vẫn được cập nhật đúng\nLog thanh toán được ghi"],
    ["F.5", "Webhook VNPay lặp", "Payment", "P0", "VNPay sandbox", "1. Thanh toán thành công\n2. Webhook gửi lần 1 → ghi log\n3. Webhook gửi lần 2 (duplicate)", "Chỉ có 1 bản ghi trong LichSuThanhToan\nKhông duplicate"],
    ["F.6", "Hoàn tiền đơn giờ - 100%", "Refund", "P0", "Đã thanh toán VNPay", "1. Hủy đơn > 12h trước giờ bắt đầu\n2. Hệ thống gọi VNPay refund API", "Refund 100% thành công\nBản ghi LichSuThanhToan với LoaiGD = refund\nGhiChu = user_cancel\nKhách nhận thông báo"],
    ["F.7", "Hoàn tiền đơn giờ - 0% (có phí)", "Refund", "P0", "Đã thanh toán VNPay", "1. Hủy đơn < 12h trước giờ bắt đầu", "Không refund\nHiển thị 'Hủy có phí'\nBản ghi phí hủy trong LichSuThanhToan"],
    ["F.8", "Hoàn tiền đơn tháng - 80%", "Refund", "P0", "Đã thanh toán gói tháng 20 buổi", "1. Đã hoàn thành 5 buổi\n2. Hủy gói\n3. Tính: 15 buổi còn lại × 80%", "Refund 80% số tiền buổi chưa làm\nLog refund với GhiChu rõ ràng"],
    ["F.9", "Ghi log refund", "Refund", "P1", "Có refund", "1. Kiểm tra bảng LichSuThanhToan sau refund", "Có bản ghi với:\n- LoaiGD = refund\n- SoTien đúng\n- GhiChu mô tả lý do"],
])

# Module G: Cancellation (7 cases)
test_cases_data.extend([
    ["G.1", "Khách hủy đơn giờ > 12h trước", "Cancellation", "P0", "Đã đặt lịch, thanh toán VNPay", "1. Vào chi tiết đơn\n2. Nhấn 'Hủy đơn'\n3. Xác nhận", "Đơn cancelled\nRefund 100%\nThông báo refund"],
    ["G.2", "Khách hủy đơn giờ < 12h trước", "Cancellation", "P0", "Đã đặt lịch, thanh toán VNPay", "1. Đặt lịch giờ 10:00 hôm nay\n2. Lúc 9:00, hủy đơn\n3. Xác nhận", "Đơn cancelled\nHiển thị 'Hủy có phí'\nKhông refund"],
    ["G.3", "Khách hủy đơn tháng", "Cancellation", "P0", "Đã đặt gói 20 buổi, làm 5 buổi", "1. Hủy gói\n2. Xác nhận", "Đơn cancelled\nRefund 80% × 15 buổi còn lại"],
    ["G.4", "Auto-cancel đơn giờ sau 2h", "Cancellation", "P0", "Đơn finding_staff, giờ bắt đầu 10:00", "1. Đến 8:00 (2h trước)\n2. AutoCancelOrdersJob chạy", "Đơn tự động cancelled\nRefund 100% (nếu VNPay)\nThông báo cho khách"],
    ["G.5", "Auto-cancel đơn đã có NV", "Cancellation", "P0", "Đơn assigned, giờ bắt đầu 10:00", "1. Đến 8:00\n2. Job chạy", "Đơn vẫn bị cancel\nRefund\nThông báo cho khách và NV"],
    ["G.6", "Admin hủy đơn", "Cancellation", "P1", "Admin login", "1. Vào 'Quản lý đơn'\n2. Chọn đơn\n3. Nhấn 'Hủy đơn'\n4. Xác nhận", "Đơn cancelled\nRefund nếu thanh toán online"],
    ["G.7", "Admin hủy 1 buổi trong gói tháng", "Cancellation", "P1", "Gói tháng có nhiều buổi", "1. Admin chọn 1 buổi\n2. Nhấn 'Hủy buổi này'", "Buổi đó cancelled\nCác buổi khác vẫn bình thường"],
])

# Module H: Notifications (10 cases)
test_cases_data.extend([
    ["H.1", "Thông báo đặt lịch thành công", "Notification", "P1", "Đặt lịch xong", "1. Khách đặt lịch thành công", "Nhận push notification\nNội dung: 'Đặt lịch thành công'\nIn-app notification"],
    ["H.2", "Thông báo gán nhân viên", "Notification", "P1", "Đơn được gán NV", "1. NV claim hoặc admin gán", "Khách nhận push\nNội dung: 'Đã tìm được nhân viên [Tên NV]'"],
    ["H.3", "Thông báo NV có đơn mới", "Notification", "P1", "Có đơn mới finding_staff", "1. Khách đặt lịch", "NV nhận push 'Có đơn mới phù hợp'\nNV mở app → thấy đơn trong 'Khả dụng'"],
    ["H.4", "Thông báo tìm NV chậm", "Notification", "P1", "Đơn finding_staff > 1/3 thời gian", "1. Đơn giờ bắt đầu 12:00 hôm nay\n2. Tạo lúc 6:00\n3. Đến 10:00 (1/3 thời gian = 2h)", "Khách nhận thông báo 'Hệ thống đang tìm nhân viên cho bạn'"],
    ["H.5", "Thông báo hủy đơn", "Notification", "P1", "Đơn bị hủy", "1. Khách hoặc admin hủy đơn", "Khách và NV nhận push 'Đơn [ID] đã bị hủy'"],
    ["H.6", "Thông báo hoàn tiền", "Notification", "P1", "Refund thành công", "1. Hủy đơn → refund", "Khách nhận push 'Hoàn tiền [số tiền] VNĐ thành công'"],
    ["H.7", "Thông báo hoàn thành", "Notification", "P1", "Đơn completed", "1. NV hoàn thành đơn", "Khách nhận push 'Đơn hoàn thành, vui lòng đánh giá'"],
    ["H.8", "Push khi app đóng", "Notification", "P1", "OneSignal config, app đóng", "1. Đóng app\n2. Tạo sự kiện (đặt lịch, gán NV, etc.)", "Vẫn nhận push notification trên device"],
    ["H.9", "Đánh dấu đã đọc", "Notification", "P1", "Có thông báo chưa đọc", "1. Vào list thông báo\n2. Nhấn vào 1 thông báo\n3. API: POST /api/notifications/{id}/mark-read", "Thông báo chuyển sang 'đã đọc'\nBadge unread count giảm"],
    ["H.10", "Unread count", "Notification", "P1", "Có thông báo chưa đọc", "1. API: GET /api/notifications/unread-count", "Trả về số lượng đúng"],
])

# Module I: Rating (4 cases)
test_cases_data.extend([
    ["I.1", "Đánh giá NV sau khi hoàn thành", "Rating", "P1", "Đơn completed", "1. Vào chi tiết đơn\n2. Chọn số sao 1-5\n3. Nhập comment\n4. Nhấn 'Gửi đánh giá'\n5. API: POST /api/bookings/{id}/rate", "Đánh giá được lưu\nĐiểm NV được cập nhật\nHiển thị 'Cảm ơn đánh giá'"],
    ["I.2", "Không cho đánh giá nếu chưa hoàn thành", "Rating", "P1", "Đơn assigned hoặc confirmed", "1. Vào chi tiết đơn", "Không hiển thị nút 'Đánh giá'"],
    ["I.3", "Tính điểm trung bình NV", "Rating", "P1", "NV có nhiều đánh giá", "1. Kiểm tra profile NV", "Điểm trung bình = (tổng điểm) / (số đánh giá)\nHiển thị đúng số sao"],
    ["I.4", "Validation đánh giá", "Rating", "P1", "Đơn completed", "1. Nhập rating = 0 hoặc 6\n2. Gửi", "Hiển thị lỗi 'Vui lòng chọn từ 1-5 sao'"],
])

# Module J: Admin (10 cases)
test_cases_data.extend([
    ["J.1", "Dashboard thống kê", "Admin", "P1", "Admin login", "1. Vào /admin/dashboard", "Hiển thị:\n- Tổng doanh thu\n- Số đơn hôm nay/tuần/tháng\n- Biểu đồ đơn theo trạng thái"],
    ["J.2", "CRUD dịch vụ - Tạo", "Admin", "P1", "Admin login", "1. Vào 'Quản lý dịch vụ'\n2. Nhấn 'Thêm'\n3. Nhập: tên, giá, số giờ\n4. Lưu", "Dịch vụ mới được tạo\nHiển thị trong danh sách"],
    ["J.3", "CRUD dịch vụ - Sửa", "Admin", "P1", "Đã có dịch vụ", "1. Nhấn 'Sửa' trên dịch vụ\n2. Đổi giá\n3. Lưu", "Giá được cập nhật"],
    ["J.4", "CRUD dịch vụ - Xóa", "Admin", "P1", "Đã có dịch vụ", "1. Nhấn 'Xóa'\n2. Xác nhận", "Dịch vụ bị xóa (hoặc soft delete)"],
    ["J.5", "CRUD voucher - Tạo", "Admin", "P1", "Admin login", "1. Vào 'Voucher'\n2. Tạo voucher: mã, %, max, ngày hết hạn\n3. Lưu", "Voucher được tạo\nKhách có thể dùng"],
    ["J.6", "Duyệt nhân viên", "Admin", "P0", "Có candidate pending", "1. Vào 'Ứng viên'\n2. Xem hồ sơ\n3. Nhấn 'Duyệt'", "Candidate → NV active\nCó thể nhận đơn"],
    ["J.7", "Khóa nhân viên", "Admin", "P1", "Có NV active", "1. Vào 'Nhân viên'\n2. Chọn NV\n3. Nhấn 'Khóa'", "Trạng thái inactive\nNV không claim được đơn"],
    ["J.8", "Xem chi tiết đơn", "Admin", "P1", "Có đơn", "1. Vào 'Quản lý đơn'\n2. Nhấn vào 1 đơn", "Hiển thị:\n- Thông tin khách\n- NV\n- Lịch sử thanh toán\n- Trạng thái"],
    ["J.9", "Gán NV thủ công", "Admin", "P0", "Đơn finding_staff, có NV available", "1. Xem chi tiết đơn\n2. Nhấn 'Gán nhân viên'\n3. Chọn NV\n4. Xác nhận", "Đơn assigned\nID_NV được cập nhật\nNV nhận thông báo"],
    ["J.10", "Xuất báo cáo Excel", "Admin", "P2", "Có đơn", "1. Vào 'Báo cáo'\n2. Chọn ngày từ - đến\n3. Nhấn 'Xuất Excel'", "File Excel được download\nChứa danh sách đơn, doanh thu"],
])

# Module K: Wallet (4 cases)
test_cases_data.extend([
    ["K.1", "Xem số dư ví", "Wallet", "P1", "Đăng nhập NV", "1. Vào 'Ví của tôi'\n2. API: GET /api/staff/wallet", "Hiển thị số dư hiện tại"],
    ["K.2", "Lịch sử giao dịch", "Wallet", "P1", "NV đã hoàn thành đơn", "1. Vào 'Lịch sử giao dịch'\n2. API: GET /api/staff/wallet/history", "Hiển thị danh sách: ngày, loại, số tiền"],
    ["K.3", "Cộng tiền sau hoàn thành", "Wallet", "P0", "Đơn được complete", "1. NV hoàn thành đơn giá 100,000 VND", "Số dư ví tăng 100,000\nGhi log trong LichSuViNhanVien"],
    ["K.4", "Báo cáo thu nhập tuần", "Wallet", "P1", "NV đã làm việc", "1. API: GET /api/staff/weekly-report", "Hiển thị tổng thu nhập từ thứ 2 - CN tuần này"],
])

# Module L: Jobs (5 cases)
test_cases_data.extend([
    ["L.1", "AutoCancelOrdersJob - đơn giờ", "Job", "P0", "Đơn finding_staff, giờ bắt đầu 10:00", "1. Đến 8:00 (2h trước)\n2. Job chạy (scheduler hoặc manual)", "Đơn cancelled\nRefund 100% nếu VNPay\nThông báo cho khách"],
    ["L.2", "AutoCancelOrdersJob - đơn tháng buổi đầu", "Job", "P0", "Gói tháng, buổi 1 finding_staff, giờ 10:00", "1. Đến 8:00\n2. Job chạy", "Buổi 1 cancelled\nCác buổi khác vẫn bình thường"],
    ["L.3", "AutoCompleteOrdersJob", "Job", "P0", "Đơn confirmed, giờ kết thúc 11:00", "1. Đến 11:00\n2. Job chạy", "Đơn tự động completed\nTiền cộng vào ví NV"],
    ["L.4", "NotifyFindingStaffDelayJob", "Job", "P1", "Đơn finding_staff, tạo lúc 6:00, giờ bắt đầu 12:00", "1. Đến 10:00 (1/3 × 6h = 2h)\n2. Job chạy", "Thông báo gửi cho khách 'Đang tìm NV'"],
    ["L.5", "Job không chạy trùng", "Job", "P0", "2 job instances", "1. Job 1 bắt đầu xử lý đơn A\n2. Job 2 chạy cùng lúc", "Chỉ 1 job xử lý đơn A\nKhông duplicate"],
])

# Module M: UI/UX (6 cases)
test_cases_data.extend([
    ["M.1", "Responsive - Desktop 1920x1080", "UI/UX", "P2", "Browser mở", "1. Truy cập web trên desktop\n2. Resize window về 1920x1080", "Giao diện hiển thị đúng, không bị vỡ"],
    ["M.2", "Responsive - Mobile 375x667", "UI/UX", "P2", "Browser mở", "1. F12 → Mobile mode\n2. Chọn iPhone SE", "Menu hamburger, layout mobile đúng"],
    ["M.3", "Flutter App - Android mượt", "UI/UX", "P1", "App cài trên Android", "1. Mở app\n2. Navigate giữacác màn hình", "Không lag, mượt mà"],
    ["M.4", "Loading states", "UI/UX", "P2", "Gọi API", "1. Đặt lịch\n2. Quan sát khi API đang xử lý", "Hiển thị loading spinner"],
    ["M.5", "Error states", "UI/UX", "P2", "API lỗi", "1. Disconnect internet\n2. Gọi API", "Hiển thị 'Lỗi kết nối, vui lòng thử lại'"],
    ["M.6", "Empty states", "UI/UX", "P2", "Chưa có đơn", "1. Vào 'Đơn của tôi'", "Hiển thị 'Bạn chưa có đơn đặt nào'"],
])

# Module N: Security (7 cases)
test_cases_data.extend([
    ["N.1", "CSRF Protection", "Security", "P0", "Web form", "1. Inspect form đăng nhập\n2. Kiểm tra có _token", "Form có CSRF token"],
    ["N.2", "SQL Injection", "Security", "P0", "Login form", "1. Nhập username: admin' OR '1'='1\n2. Submit", "Không đăng nhập được\nInput bị escape"],
    ["N.3", "XSS", "Security", "P0", "Comment field", "1. Nhập: <script>alert('XSS')</script>\n2. Submit", "Script không chạy\nBị escape thành text"],
    ["N.4", "Password hashing", "Security", "P0", "Database", "1. Đăng ký user\n2. Kiểm tra bảng TaiKhoan", "Password lưu dạng hash (bcrypt), không plain text"],
    ["N.5", "Authorization - IDOR", "Security", "P0", "User A login", "1. User A lấy ID đơn của User B\n2. Truy cập /api/bookings/{id_of_B}", "Trả về 403 Forbidden hoặc không hiển thị"],
    ["N.6", "File upload - extension", "Security", "P1", "Upload form", "1. Upload file .exe\n2. Submit", "Hiển thị lỗi 'Chỉ chấp nhận .jpg, .png'"],
    ["N.7", "Rate limiting", "Security", "P1", "API endpoint", "1. Gọi API login 100 lần trong 1 phút", "Sau 10-20 lần → 'Too many requests'"],
])

# Module O: Performance (5 cases)
test_cases_data.extend([
    ["O.1", "API response time - đặt lịch", "Performance", "P1", "Server chạy", "1. POST /api/bookings\n2. Đo thời gian response", "< 1s"],
    ["O.2", "API response time - danh sách", "Performance", "P1", "Có 100 đơn", "1. GET /api/bookings\n2. Đo thời gian", "< 500ms"],
    ["O.3", "Database query - N+1", "Performance", "P1", "Database có data", "1. Enable query log\n2. Gọi API danh sách đơn\n3. Kiểm tra log", "Không có N+1 query (dùng eager loading)"],
    ["O.4", "Pagination", "Performance", "P1", "Có 200 đơn", "1. GET /api/bookings", "Có pagination, mỗi page 20-50 items"],
    ["O.5", "Load test - 50 users", "Performance", "P2", "k6/JMeter setup", "1. Chạy load test 50 concurrent users\n2. Mỗi user đặt lịch", "Hệ thống vẫn hoạt động\nKhông crash\nResponse time < 3s"],
])

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Headers
headers = ["ID", "Tên Test Case", "Module", "Priority", "Điều kiện tiên quyết", "Các bước thực hiện", "Kết quả mong đợi", "Trạng thái", "Người test", "Ngày test", "Ghi chú"]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 50
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 25

# Write data
row_num = 2
for test_case in test_cases_data:
    for col_num, value in enumerate(test_case, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Color coding for priority
        if col_num == 4:
            if value == "P0":
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                cell.font = Font(bold=True, color="CC0000")
            elif value == "P1":
                cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
                cell.font = Font(bold=True, color="FF8C00")
            elif value == "P2":
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                cell.font = Font(color="0066CC")
    
    row_num += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:K{row_num-1}"

# Save file
output_path = r"D:\Web_Cung_Cap_Dich_Vu_Giup_Viec_Theo_Gio\Test_Cases_He_Thong_Giup_Viec.xlsx"
wb.save(output_path)

print(f"✅ File Excel đã được tạo thành công!")
print(f"📁 Đường dẫn: {output_path}")
print(f"📊 Tổng số test cases: {len(test_cases_data)}")
print(f"")
print(f"Phân bố theo module:")
modules = {}
for tc in test_cases_data:
    module = tc[2]
    modules[module] = modules.get(module, 0) + 1
for module, count in sorted(modules.items()):
    print(f"  - {module}: {count} test cases")
